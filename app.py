import os
import openpyxl

from flask import Flask, render_template, make_response

app = Flask(__name__, static_url_path='/static')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/output_excel')
def output_excel():
    filename = "sample.xlsx"

    wbook = openpyxl.Workbook()
    wsheet = wbook.active
    wsheet.cell(2, 2).value = "test"

    fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='00FFFF', bgColor='00FFFF')
    wsheet["B2"].fill = fill

    side = openpyxl.styles.borders.Side(style='thin', color='000000')
    border = openpyxl.styles.borders.Border(top=side, bottom=side, left=side, right=side)
    wsheet["D2"].border = border
    wsheet["F2"].border = border
    for row in wsheet["C11:D13"]:
        for cell in row:
            wsheet[cell.coordinate].border = border

    font = openpyxl.styles.Font(name='メイリオ', size=20, bold=True)
    wsheet["B4"].value = "test"
    wsheet["B4"].font = font

    wsheet.column_dimensions["B"].width = 50

    wsheet.merge_cells("B6:C8")
    wsheet["B6"].value = "test"
    wsheet["B6"].border = border
    for cell in wsheet["B6:C8"][0]:
        cell.alignment = openpyxl.styles.Alignment(vertical="center")

    wsheet.cell(16, 2).value = 1234567
    wsheet.cell(16, 2).number_format = '#,##0'

    wbook.save(filename)
    wbook.close()

    xlsx_mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    res = make_response()

    with open(filename, "rb") as wb:
        res.data = wb.read()

    res.headers["Content-Disposition"] = "attachment; filename=" + filename
    res.mimetype = xlsx_mimetype
    os.remove(filename)

    return res


if __name__ == '__main__':
    app.run(debug=True)
